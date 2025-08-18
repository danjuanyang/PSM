# PSM/app/activity/routes.py
from flask import request, session, jsonify, make_response, current_app
from flask_login import current_user, login_required
from datetime import datetime, date, timedelta
import io
import openpyxl


from . import activity_bp
from ..decorators import permission_required
from ..models import db, UserSession, UserActivityLog, User, RoleEnum, Project, Subproject, ProjectStage, StageTask, UserEntityActivity


@activity_bp.route('/heartbeat', methods=['POST'])
@login_required
def heartbeat():
    """
    接收前端的心跳请求，更新会话的最后活动时间，并记录活动日志。
    此版本更健壮，不依赖于Flask session中的特定ID，而是依赖于当前登录用户。
    """
    # @login_required 确保了 current_user 是有效的。
    # 我们查找该用户最新的、仍处于活动状态的会话记录。
    user_session = UserSession.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).order_by(UserSession.login_time.desc()).first()

    if not user_session:
        # 如果用户已登录但没有活动的会话记录（例如，数据库被手动清理），
        # 返回一个错误强制前端重新登录，以同步状态。
        return jsonify({"status": "error", "message": "活动会话记录未找到，请重新登录"}), 401

    # 更新心跳时间
    user_session.last_activity_time = datetime.now()

    # 记录活动日志 (使用 silent=True 使JSON解析更安全)
    data = request.get_json(silent=True) or {}
    module = data.get('module')

    activity_log = UserActivityLog(
        user_id=current_user.id,
        session_id=user_session.id,  # 使用我们刚刚找到的会话ID
        action_type='HEARTBEAT',
        endpoint=request.endpoint,
        module=module,
        status_code=200,
        ip_address=request.remote_addr
    )
    db.session.add(activity_log)

    try:
        db.session.commit()
        return jsonify({"status": "ok"}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"心跳错误： {e}")
        return jsonify({"status": "error", "message": "数据库操作失败"}), 500


@activity_bp.route('/unload', methods=['POST'])
@login_required
def unload():
    """
    接收前端在页面卸载时发送的信号，以正常结束会话。
    """
    session_id = session.get('user_session_id')
    if not session_id:
        return jsonify({"status": "ok", "message": "会话已结束"}), 200

    user_session = UserSession.query.get(session_id)
    if user_session and user_session.is_active:
        user_session.is_active = False
        user_session.logout_time = datetime.now()
        if user_session.login_time:
            duration = user_session.logout_time - user_session.login_time
            user_session.session_duration = int(duration.total_seconds())

        try:
            db.session.commit()
            # 从会话中移除，以防万一
            session.pop('user_session_id', None)
        except Exception as e:
            db.session.rollback()
            print(f"Error in unload: {e}")
            return jsonify({"status": "error", "message": "数据库操作失败"}), 500

    return jsonify({"status": "ok"}), 200


@activity_bp.route('/stats', methods=['GET'])
@permission_required('view_user_activity')
@login_required
def get_activity_stats():
    """
    获取用户活动统计数据，主要是总在线时长。
    """
    # 查询每个用户的总会话时长
    user_stats = db.session.query(
        User.username,
        db.func.sum(UserSession.session_duration).label('total_duration')
    ).join(UserSession, User.id == UserSession.user_id).filter(UserSession.session_duration.isnot(None)).group_by(
        User.username).order_by(db.desc('total_duration')).filter(User.role != RoleEnum.SUPER)

    # 格式化结果
    stats_data = [
        {'username': username, 'total_duration': total_duration}
        for username, total_duration in user_stats
    ]

    return jsonify(stats_data)


def _query_module_stats_data(is_export=False):
    """
    Helper function to reuse module stats query logic.
    Handles both JSON response and data for export.
    """
    from sqlalchemy import func, desc, cast, Date

    user_ids_str = request.args.get('user_ids')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    base_query = db.session.query(
        UserActivityLog.module,
        UserActivityLog.user_id,
        UserActivityLog.timestamp
    ).filter(UserActivityLog.module.isnot(None))

    if user_ids_str:
        try:
            user_ids = [int(uid) for uid in user_ids_str.split(',')]
            if user_ids:
                base_query = base_query.filter(UserActivityLog.user_id.in_(user_ids))
        except ValueError:
            if not is_export: return jsonify({"status": "error", "message": "无效的用户ID格式"}), 400
            else: return []

    # --- FIX: Correct date filtering ---
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            base_query = base_query.filter(db.func.date(UserActivityLog.timestamp) >= start_date)
        except ValueError:
            if not is_export: return jsonify({"status": "error", "message": "无效的开始日期格式"}), 400
            else: return []
            
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            base_query = base_query.filter(db.func.date(UserActivityLog.timestamp) <= end_date)
        except ValueError:
            if not is_export: return jsonify({"status": "error", "message": "无效的结束日期格式"}), 400
            else: return []

    lead_func = func.lead(UserActivityLog.timestamp, 1).over(
        partition_by=(UserActivityLog.user_id, func.date(UserActivityLog.timestamp)),
        order_by=UserActivityLog.timestamp
    )
    duration_calc = (func.julianday(lead_func) - func.julianday(UserActivityLog.timestamp)) * 86400.0
    activity_with_duration = base_query.add_columns(duration_calc.label('duration')).subquery()

    module_stats = db.session.query(
        activity_with_duration.c.module,
        func.sum(activity_with_duration.c.duration).label('total_duration_seconds')
    ).filter(
        activity_with_duration.c.duration.isnot(None),
        activity_with_duration.c.duration > 0,
        activity_with_duration.c.duration < 1800
    ).group_by(activity_with_duration.c.module).order_by(desc('total_duration_seconds')).all()
    
    return module_stats

@activity_bp.route('/module_stats', methods=['GET'])
@permission_required('view_user_activity')
@login_required
def get_module_activity_stats():
    """
    计算并返回每个模块的用户总停留时间。
    """
    result = _query_module_stats_data()
    if isinstance(result, tuple): # Error case
        return result

    stats_data = [
        {'module': module, 'duration_seconds': int(total_duration) if total_duration else 0}
        for module, total_duration in result
    ]
    return jsonify(stats_data)


@activity_bp.route('/entity_stats', methods=['GET'])
@login_required
@permission_required('view_user_activity')
def get_entity_activity_stats():
    entity_type = request.args.get('entity_type')
    entity_id = request.args.get('entity_id', type=int)
    if not entity_type or not entity_id: return jsonify({"status": "error", "message": "必须提供 entity_type 和 entity_id"}), 400
    entity_model_map = {'project': Project, 'subproject': Subproject, 'stage': ProjectStage, 'task': StageTask}
    entity_model = entity_model_map.get(entity_type)
    if not entity_model: return jsonify({"status": "error", "message": f"无效的 entity_type: {entity_type}"}), 400
    entity = db.session.get(entity_model, entity_id)
    if not entity: return jsonify({"status": "error", "message": "实体未找到"}), 404
    total_stats = {'entity_name': getattr(entity, 'name', 'N/A'), 'edit_count': entity.edit_count, 'total_edit_duration_seconds': entity.total_edit_duration}
    activity_logs = db.session.query(UserEntityActivity, User.username).join(User, UserEntityActivity.user_id == User.id).filter(UserEntityActivity.entity_type == entity_type, UserEntityActivity.entity_id == entity_id).order_by(UserEntityActivity.created_at.desc()).all()
    detailed_logs = [{'username': username, 'duration_seconds': log.duration_seconds, 'edited_at': log.created_at.isoformat()} for log, username in activity_logs]
    return jsonify({"status": "ok", "total_stats": total_stats, "detailed_logs": detailed_logs})

def get_entity_type_from_class(entity_class):
    """Helper to map model class to simple string type."""
    if entity_class == Project:
        return 'project'
    if entity_class == Subproject:
        return 'subproject'
    if entity_class == ProjectStage:
        return 'stage'
    if entity_class == StageTask:
        return 'task'
    return entity_class.__name__.lower()

def build_entity_dict(entity):
    """
    递归辅助函数，用于构建项目及其子项的统计数据树。
    """
    # --- FIX: Use helper function for correct type mapping ---
    entity_type = get_entity_type_from_class(entity.__class__)
    
    result = {
        'key': f"{entity_type}-{entity.id}",
        'entity_type': entity_type,
        'id': entity.id,
        'name': entity.name,
        'edit_count': entity.edit_count or 0,
        'total_edit_duration': entity.total_edit_duration or 0,
        'children': []
    }

    if isinstance(entity, Project):
        for subproject in entity.subprojects: result['children'].append(build_entity_dict(subproject))
    elif isinstance(entity, Subproject):
        for stage in entity.stages: result['children'].append(build_entity_dict(stage))
    elif isinstance(entity, ProjectStage):
        for task in entity.tasks: result['children'].append(build_entity_dict(task))
    
    if not result['children']:
        del result['children']
        
    return result

@activity_bp.route('/project_summary_stats', methods=['GET'])
@login_required
@permission_required('view_user_activity')
def get_project_summary_stats():
    project_id = request.args.get('project_id', type=int)
    if not project_id: return jsonify({"status": "error", "message": "必须提供 project_id"}), 400
    project = Project.query.get(project_id)
    if not project: return jsonify({"status": "error", "message": "项目未找到"}), 404
    tree_data = build_entity_dict(project)
    return jsonify({"status": "ok", "stats": tree_data})

@activity_bp.route('/user_summary_stats', methods=['GET'])
@login_required
@permission_required('view_user_activity')
def get_user_summary_stats():
    """
    获取特定用户的所有编辑活动，并附加上下文信息。
    """
    user_id = request.args.get('user_id', type=int)
    if not user_id: return jsonify({"status": "error", "message": "必须提供 user_id"}), 400
    user = User.query.get(user_id)
    if not user: return jsonify({"status": "error", "message": "用户未找到"}), 404

    activities = UserEntityActivity.query.filter_by(user_id=user_id).order_by(UserEntityActivity.created_at.desc()).all()

    # --- FIX: Preload all entities for all activity types ---
    task_ids = [a.entity_id for a in activities if a.entity_type == 'task']
    stage_ids = [a.entity_id for a in activities if a.entity_type == 'stage']
    subproject_ids = [a.entity_id for a in activities if a.entity_type == 'subproject']
    project_ids = [a.entity_id for a in activities if a.entity_type == 'project']

    tasks = {t.id: t for t in StageTask.query.filter(StageTask.id.in_(task_ids)).all()}
    stages = {s.id: s for s in ProjectStage.query.filter(ProjectStage.id.in_(stage_ids)).all()}
    subprojects = {sp.id: sp for sp in Subproject.query.filter(Subproject.id.in_(subproject_ids)).all()}
    projects = {p.id: p for p in Project.query.filter(Project.id.in_(project_ids)).all()}

    results = []
    for activity in activities:
        context = {
            'project': '-', 'subproject': '-', 'stage': '-', 'task': '-',
            'entity_type': activity.entity_type,
            'duration_seconds': activity.duration_seconds,
            'edit_count': 1,
            'edited_at': activity.created_at.isoformat()
        }
        
        # --- FIX: Complete context-aware logic for all types ---
        if activity.entity_type == 'task':
            task = tasks.get(activity.entity_id)
            if task:
                context['task'] = task.name
                if task.stage:
                    context['stage'] = task.stage.name
                    if task.stage.subproject:
                        context['subproject'] = task.stage.subproject.name
                        if task.stage.subproject.project:
                            context['project'] = task.stage.subproject.project.name
        elif activity.entity_type == 'stage':
            stage = stages.get(activity.entity_id)
            if stage:
                context['stage'] = stage.name
                if stage.subproject:
                    context['subproject'] = stage.subproject.name
                    if stage.subproject.project:
                        context['project'] = stage.subproject.project.name
        elif activity.entity_type == 'subproject':
            subproject = subprojects.get(activity.entity_id)
            if subproject:
                context['subproject'] = subproject.name
                if subproject.project:
                    context['project'] = subproject.project.name
        elif activity.entity_type == 'project':
            project = projects.get(activity.entity_id)
            if project:
                context['project'] = project.name
        
        results.append(context)

    return jsonify({"status": "ok", "stats": results})


# --- EXPORT ENDPOINTS (XLSX) ---

@activity_bp.route('/export_module_stats', methods=['GET'])
@login_required
@permission_required('view_user_activity')
def export_module_stats():
    stats_data = _query_module_stats_data(is_export=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模块停留时间"
    ws.append(['模块', '总停留时长 (秒)'])
    for row in stats_data:
        ws.append([row.module, int(row.total_duration_seconds) if row.total_duration_seconds else 0])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=module_stats.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

def _flatten_project_tree(node, level=0, flat_list=None):
    if flat_list is None: flat_list = []
    prefix = "    " * level
    flat_list.append({'name': prefix + node['name'], 'entity_type': {'project': '项目', 'subproject': '子项目', 'stage': '阶段', 'task': '任务'}.get(node['entity_type'], ''), 'edit_count': node['edit_count'], 'total_edit_duration': node['total_edit_duration']})
    if 'children' in node:
        for child in node['children']: _flatten_project_tree(child, level + 1, flat_list)
    return flat_list

@activity_bp.route('/export_project_summary', methods=['GET'])
@login_required
@permission_required('view_user_activity')
def export_project_summary():
    project_id = request.args.get('project_id', type=int)
    if not project_id: return "Error: Missing project_id", 400
    project = Project.query.get(project_id)
    if not project: return "Error: Project not found", 404
    tree_data = build_entity_dict(project)
    flat_data = _flatten_project_tree(tree_data)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目活动摘要"
    ws.append(['层级名称', '类型', '总编辑次数', '总编辑时长 (秒)'])
    for row in flat_data:
        ws.append([row['name'], row['entity_type'], row['edit_count'], row['total_edit_duration']])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=project_summary.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@activity_bp.route('/export_user_summary', methods=['GET'])
@login_required
@permission_required('view_user_activity')
def export_user_summary():
    user_id = request.args.get('user_id', type=int)
    if not user_id: return "Error: Missing user_id", 400
    response = get_user_summary_stats()
    if response.status_code != 200: return "Error: Could not retrieve user stats", response.status_code
    stats_data = response.get_json().get('stats', [])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户活动摘要"
    ws.append(['项目', '子项目', '阶段', '任务', '编辑时长 (秒)', '编辑时间'])
    for row in stats_data:
        ws.append([row['project'], row['subproject'], row['stage'], row['task'], row['duration_seconds'], row['edited_at']])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=user_summary.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response
